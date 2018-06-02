# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText 

def get_question_list():
    wb = load_workbook('questions.xlsx')
    sheet = wb.get_sheet_by_name('Определение сложности 2017');

    questions_list = {};
    check = True;
    row = 0;

    while(check == True):
        row+=1
        #print(row)
        if sheet['A'+str(row)].value!="END":
            questions_list.update({sheet['A'+str(row)].value:sheet['B'+str(row)].value});
            #print(sheet['A'+str(row)].value, sheet['B'+str(row)].value)
        else:
            check = False;
    
    return questions_list;
    #print(questions_list.items())
     

def check_difficult(score, form, for_who):
    wb = load_workbook('difference.xlsx')
    sheet = wb.get_sheet_by_name('Лист1');
    score = float(score)
    form = int(form);
    
    print(score, form, for_who)
    
    diff_list = {"name":"", "w_nds":"", "nds":""};
    
    if score<=1:
        if form==1:#P form
            if for_who=="На одного":
                diff_list.update({"name":sheet['B1'].value})
                diff_list.update({"w_nds":sheet['C1'].value})
                diff_list.update({"nds":int(sheet['С1'].value)*1.21})
            if for_who=="На семью":
                diff_list.update({"name":sheet['E1'].value})
                diff_list.update({"w_nds":sheet['F1'].value})
                diff_list.update({"nds":int(sheet['F1'].value)*1.21})
        if form==2:#M form
            if for_who=="На одного":
                diff_list.update({"name":sheet['B6'].value})
                diff_list.update({"w_nds":sheet['C6'].value})
                diff_list.update({"nds":int(sheet['C6'].value)*1.21})
            if for_who=="На семью":
                diff_list.update({"name":sheet['E6'].value})
                diff_list.update({"w_nds":sheet['F6'].value})
                diff_list.update({"nds":int(sheet['F6'].value)*1.21})
        if form==3:#C form
            if for_who=="На одного":
                diff_list.update({"name":sheet['B11'].value})
                diff_list.update({"w_nds":sheet['C11'].value})
                diff_list.update({"nds":int(sheet['C11'].value)*1.21})
            if for_who=="На семью":
                diff_list.update({"name":sheet['E11'].value})
                diff_list.update({"w_nds":sheet['F11'].value})
                diff_list.update({"nds":int(sheet['F11'].value)*1.21})
    elif score>1 and score<=2:
        if form==1:#P form
            if for_who=="На одного":
                diff_list.update({"name":sheet['B2'].value})
                diff_list.update({"w_nds":sheet['C2'].value})
                diff_list.update({"nds":int(sheet['C2'].value)*1.21})
            if for_who=="На семью":
                diff_list.update({"name":sheet['E2'].value})
                diff_list.update({"w_nds":sheet['F2'].value})
                diff_list.update({"nds":int(sheet['F2'].value)*1.21})
        if form==2:#M form
            if for_who=="На одного":
                diff_list.update({"name":sheet['B7'].value})
                diff_list.update({"w_nds":sheet['C7'].value})
                diff_list.update({"nds":int(sheet['C7'].value)*1.21})
            if for_who=="На семью":
                diff_list.update({"name":sheet['E7'].value})
                diff_list.update({"w_nds":sheet['F7'].value})
                diff_list.update({"nds":int(sheet['F7'].value)*1.21})
        if form==3:#C form
            if for_who=="На одного":
                diff_list.update({"name":sheet['B12'].value})
                diff_list.update({"w_nds":sheet['C12'].value})
                diff_list.update({"nds":int(sheet['C12'].value)*1.21})
            if for_who=="На семью":
                diff_list.update({"name":sheet['E12'].value})
                diff_list.update({"w_nds":sheet['F12'].value})
                diff_list.update({"nds":int(sheet['F12'].value)*1.21})
    elif score>2 and score<=4.5:
        if form==1:#P form
            if for_who=="На одного":
                diff_list.update({"name":sheet['B3'].value})
                diff_list.update({"w_nds":sheet['C3'].value})
                diff_list.update({"nds":int(sheet['C3'].value)*1.21})
            if for_who=="На семью":
                diff_list.update({"name":sheet['E3'].value})
                diff_list.update({"w_nds":sheet['F3'].value})
                diff_list.update({"nds":int(sheet['F3'].value)*1.21})
        if form==2:#M form
            if for_who=="На одного":
                diff_list.update({"name":sheet['B8'].value})
                diff_list.update({"w_nds":sheet['C8'].value})
                diff_list.update({"nds":int(sheet['C8'].value)*1.21})
            if for_who=="На семью":
                diff_list.update({"name":sheet['E8'].value})
                diff_list.update({"w_nds":sheet['F8'].value})
                diff_list.update({"nds":int(sheet['F8'].value)*1.21})
        if form==3:#C form
            if for_who=="На одного":
                diff_list.update({"name":sheet['B13'].value})
                diff_list.update({"w_nds":sheet['C13'].value})
                diff_list.update({"nds":int(sheet['C13'].value)*1.21})
            if for_who=="На семью":
                diff_list.update({"name":sheet['E13'].value})
                diff_list.update({"w_nds":sheet['F13'].value})
                diff_list.update({"nds":int(sheet['F13'].value)*1.21})
    elif score>4.5:
        if form==1:#P form
            if for_who=="На одного":
                diff_list.update({"name":sheet['B4'].value})
                diff_list.update({"w_nds":sheet['C4'].value})
                diff_list.update({"nds":int(sheet['C4'].value)*1.21})
            if for_who=="На семью":
                diff_list.update({"name":sheet['E4'].value})
                diff_list.update({"w_nds":sheet['F4'].value})
                diff_list.update({"nds":int(sheet['F4'].value)*1.21})
        if form==2:#M form
            if for_who=="На одного":
                diff_list.update({"name":sheet['B9'].value})
                diff_list.update({"w_nds":sheet['C9'].value})
                diff_list.update({"nds":int(sheet['C9'].value)*1.21})
            if for_who=="На семью":
                diff_list.update({"name":sheet['E9'].value})
                diff_list.update({"w_nds":sheet['F9'].value})
                diff_list.update({"nds":int(sheet['F9'].value)*1.21})
        if form==3:#C form
            if for_who=="На одного":
                diff_list.update({"name":sheet['B14'].value})
                diff_list.update({"w_nds":sheet['C14'].value})
                diff_list.update({"nds":int(sheet['C14'].value)*1.21})
            if for_who=="На семью":
                diff_list.update({"name":sheet['E14'].value})
                diff_list.update({"w_nds":sheet['F14'].value})
                diff_list.update({"nds":int(sheet['F14'].value)*1.21})
    
    return diff_list;
            

#print(check_difficult(0.5, 3, "На одного"))

""" Сложность 1 = 0   to 1
    Сложность 2	= 1,1 to 2
    Сложность 3	= 2,1 to 4,5
    Сложность 4	= 4,6 to 30"""

def send_all(text, client_email):
    to = "info@nalog.nl";
    
    s  = smtplib.SMTP('smtp.gmail.com', 587);

    s.starttls()
    
    me_login = 'nalogalert@gmail.com'
    me_pass = 'Hesoyam1q3e5t2w4r'
    
    s.login(me_login, me_pass)
    
            
    msg = MIMEText(text)
    msg['Subject'] = '[Telegram] Декларация заказ от '+client_email
    msg['From'] = me_login
    msg['To'] = to
            
    s.sendmail(me_login , to, msg.as_string())
    s.quit();

#send_all("213")
